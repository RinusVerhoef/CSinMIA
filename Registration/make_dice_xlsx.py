import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import statistics

src = openpyxl.load_workbook(
    "C:/Users/30697/OneDrive/2.Netherlands/capita_results/patient_comparison_affine5_affine20_affine64_nonlinear.xlsx",
    read_only=True,
)
ws_src = src.active
rows = list(ws_src.iter_rows(values_only=True))
header = rows[0]

# Column indices
cols = {}
for prefix, name in [("affine5", "a5"), ("affine20", "a20"), ("affine64", "a64"), ("nonlinear", "nl")]:
    for metric in ["Dice", "Jaccard", "HD_mm", "RVD", "Time_s"]:
        key = f"{name}_{metric.lower()}"
        col_name = f"{prefix}_{metric}"
        if col_name in header:
            cols[key] = header.index(col_name)

data = []
for row in rows[1:]:
    d = {"pat": row[0]}
    for prefix in ["a5", "a20", "a64", "nl"]:
        for metric in ["dice", "jaccard", "hd_mm", "rvd", "time_s"]:
            key = f"{prefix}_{metric}"
            if key in cols:
                d[key] = row[cols[key]] if row[cols[key]] else 0
            else:
                d[key] = 0
    data.append(d)
src.close()

n = len(data)
wb = openpyxl.Workbook()

hdr_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2F5496")
hdr_align = Alignment(horizontal="center", vertical="center")
num_font = Font(name="Arial", size=10)
green_fill = PatternFill("solid", fgColor="C6EFCE")
red_fill = PatternFill("solid", fgColor="FFC7CE")
summary_fill = PatternFill("solid", fgColor="D6DCE4")
summary_font = Font(name="Arial", bold=True, size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def style_header(ws, cols_count):
    for c in range(1, cols_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

def add_metric_sheet(wb, title, metric_key, num_format, higher_is_better=True, first=False):
    if first:
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title)

    ws.append(["Patient", "Affine 5", "Affine 20", "Affine 64", "NonLinear", "Best Method"])
    style_header(ws, 6)

    labels = ["Affine 5", "Affine 20", "Affine 64", "NonLinear"]
    prefixes = ["a5", "a20", "a64", "nl"]
    wins = {l: 0 for l in labels}

    for d in data:
        vals = [d[f"{p}_{metric_key}"] for p in prefixes]
        if higher_is_better:
            best_val = max(vals)
        else:
            positive = [v for v in vals if v > 0]
            best_val = min(positive) if positive else 0
        best_method = labels[vals.index(best_val)] if best_val != 0 else "N/A"
        if best_method != "N/A":
            wins[best_method] += 1
        ws.append([d["pat"]] + vals + [best_method])

    for r in range(2, n + 2):
        vals = [ws.cell(row=r, column=c).value for c in range(2, 6)]
        if higher_is_better:
            best_v = max(vals)
            positive = [v for v in vals if v > 0]
            worst_v = min(positive) if positive else 0
        else:
            positive = [v for v in vals if v > 0]
            best_v = min(positive) if positive else 0
            worst_v = max(vals)

        for c in range(2, 6):
            cell = ws.cell(row=r, column=c)
            cell.number_format = num_format
            cell.font = num_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            if cell.value == best_v and cell.value != 0:
                cell.fill = green_fill
            elif cell.value == worst_v and cell.value != 0:
                cell.fill = red_fill
        ws.cell(row=r, column=1).font = num_font
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=6).font = num_font
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=6).border = thin_border

    sr = n + 2
    for label, formula_fn in [("Mean", "AVERAGE"), ("Median", "MEDIAN"), ("Min", "MIN"), ("Max", "MAX")]:
        ws.cell(row=sr, column=1, value=label).font = summary_font
        ws.cell(row=sr, column=1).fill = summary_fill
        ws.cell(row=sr, column=1).border = thin_border
        ws.cell(row=sr, column=1).alignment = Alignment(horizontal="center")
        for c in range(2, 6):
            col_letter = get_column_letter(c)
            ws.cell(row=sr, column=c, value=f"={formula_fn}({col_letter}2:{col_letter}{n+1})")
            ws.cell(row=sr, column=c).number_format = num_format
            ws.cell(row=sr, column=c).font = summary_font
            ws.cell(row=sr, column=c).fill = summary_fill
            ws.cell(row=sr, column=c).border = thin_border
            ws.cell(row=sr, column=c).alignment = Alignment(horizontal="center")
        sr += 1

    ws.cell(row=sr, column=1, value="Wins").font = summary_font
    ws.cell(row=sr, column=1).fill = summary_fill
    ws.cell(row=sr, column=1).border = thin_border
    ws.cell(row=sr, column=1).alignment = Alignment(horizontal="center")
    for c, method in zip(range(2, 6), labels):
        ws.cell(row=sr, column=c, value=wins[method])
        ws.cell(row=sr, column=c).font = summary_font
        ws.cell(row=sr, column=c).fill = summary_fill
        ws.cell(row=sr, column=c).border = thin_border
        ws.cell(row=sr, column=c).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 10
    for c in range(2, 7):
        ws.column_dimensions[get_column_letter(c)].width = 14
    return ws

# Create all metric sheets
add_metric_sheet(wb, "Dice", "dice", "0.0000", higher_is_better=True, first=True)
add_metric_sheet(wb, "Jaccard", "jaccard", "0.0000", higher_is_better=True)
add_metric_sheet(wb, "Hausdorff (mm)", "hd_mm", "0.00", higher_is_better=False)
add_metric_sheet(wb, "RVD", "rvd", "0.0000", higher_is_better=False)

# Time sheet (lower is better)
add_metric_sheet(wb, "Time (s)", "time_s", "0.0", higher_is_better=False)

# ==================== Summary Sheet ====================
ws_sum = wb.create_sheet("Summary")
ws_sum.append(["Metric", "Affine 5", "Affine 20", "Affine 64", "NonLinear"])
style_header(ws_sum, 5)

prefixes = ["a5", "a20", "a64", "nl"]
method_names = ["Affine 5", "Affine 20", "Affine 64", "NonLinear"]

def safe_mean(vals):
    pos = [v for v in vals if v > 0]
    return round(statistics.mean(pos), 4) if pos else 0

summary_data = [
    ("Mean Dice", [safe_mean([d[f"{p}_dice"] for d in data]) for p in prefixes], "0.0000"),
    ("Median Dice", [round(statistics.median([d[f"{p}_dice"] for d in data]), 4) for p in prefixes], "0.0000"),
    ("Mean Jaccard", [safe_mean([d[f"{p}_jaccard"] for d in data]) for p in prefixes], "0.0000"),
    ("Mean HD (mm)", [safe_mean([d[f"{p}_hd_mm"] for d in data]) for p in prefixes], "0.00"),
    ("Mean |RVD|", [round(statistics.mean([abs(d[f"{p}_rvd"]) for d in data]), 4) for p in prefixes], "0.0000"),
    ("Mean Time (s)", [safe_mean([d[f"{p}_time_s"] for d in data]) for p in prefixes], "0.0"),
]

for label, vals, fmt in summary_data:
    ws_sum.append([label] + vals)

for r in range(2, len(summary_data) + 2):
    for c in range(1, 6):
        cell = ws_sum.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        cell.font = num_font
        if c > 1:
            cell.number_format = summary_data[r-2][2]
    ws_sum.cell(row=r, column=1).font = summary_font

ws_sum.column_dimensions["A"].width = 16
for c in range(2, 6):
    ws_sum.column_dimensions[get_column_letter(c)].width = 14

out_path = "C:/Users/30697/OneDrive/2.Netherlands/capita_results/all_metrics_comparison.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Patients: {n}")
print(f"Sheets: {wb.sheetnames}")
